"""
attendance.py - Slot-based attendance logic.

Rules:
  - Present: student detected by camera during slot
  - Absent:  auto-marked at slot end for students not detected
  - SlotWatcher thread monitors slot transitions and triggers auto-absent
  - No Late concept — only Present or Absent
"""

import os, csv, logging, threading, time
from datetime import datetime, date, timedelta
import database

logger = logging.getLogger(__name__)
ATTENDANCE_DIR = "attendance_records"

# ── Cooldown tracker ───────────────────────────────────────────────────────────
_cooldown_tracker: dict = {}

def _ensure_dir(): os.makedirs(ATTENDANCE_DIR, exist_ok=True)

def _get_cooldown() -> int:
    return int(database.get_setting("cooldown_minutes", "50"))

def _is_on_cooldown(student_id, slot_label) -> bool:
    last = _cooldown_tracker.get((student_id, slot_label))
    if last is None: return False
    return (datetime.now() - last) < timedelta(minutes=_get_cooldown())

def _set_cooldown(student_id, slot_label):
    _cooldown_tracker[(student_id, slot_label)] = datetime.now()

def reset_cooldowns():
    _cooldown_tracker.clear()
    logger.info("Cooldowns reset.")


# ── Active session helper ──────────────────────────────────────────────────────

def get_active_session(now=None):
    """Return (slot_label, 'Present') if active, else (None, None)."""
    slot = database.get_current_slot(now)
    return (slot["label"], "Present") if slot else (None, None)


# ── Mark Present (called by camera) ───────────────────────────────────────────

def mark_attendance(student_id: str, name: str) -> tuple[bool, str]:
    """
    Mark student Present in current slot.
    Silently ignored if:
      - No active slot
      - Cooldown active (already processed recently)
      - Already marked for this slot today
    """
    now     = datetime.now()
    today   = now.strftime("%Y-%m-%d")
    now_str = now.strftime("%H:%M:%S")

    slot = database.get_current_slot(now)
    if slot is None:
        return False, f"No active slot."

    slot_label = slot["label"]

    if _is_on_cooldown(student_id, slot_label):
        return False, f"Cooldown active for {name}."

    _set_cooldown(student_id, slot_label)

    saved = database.insert_attendance(
        student_id, name, today, now_str, slot_label, "auto", "Present")

    if not saved:
        # Check if it was Absent and upgrade to Present
        conn = database.get_connection(); c = conn.cursor()
        c.execute("SELECT status FROM attendance WHERE student_id=? AND date=? AND slot=?",
                  (student_id, today, slot_label))
        existing = c.fetchone(); conn.close()
        if existing and existing["status"] == "Absent":
            conn = database.get_connection(); c = conn.cursor()
            c.execute("UPDATE attendance SET status='Present', marked_by='auto', time=? "
                      "WHERE student_id=? AND date=? AND slot=?",
                      (now_str, student_id, today, slot_label))
            conn.commit(); conn.close()
            _write_csv(student_id, name, today, now_str, slot_label, "Present")
            msg = f"{name} updated Present — {slot_label}"
            logger.info(msg)
            return True, msg
        return False, f"{name} already marked Present for {slot_label}."

    _write_csv(student_id, name, today, now_str, slot_label, "Present")
    msg = f"{name} marked Present — {slot_label} @ {now_str}"
    logger.info(msg)
    return True, msg


# ── Auto-absent watcher ────────────────────────────────────────────────────────

_watcher_running  = False
_watcher_thread   = None
_last_slot_seen   = None   # tracks which slot was last active


def start_slot_watcher():
    """
    Start the background thread that watches slot transitions.
    When a slot ends → auto-mark Absent for all students not detected.
    Safe to call multiple times — only starts one thread.
    """
    global _watcher_running, _watcher_thread, _last_slot_seen
    if _watcher_running:
        return
    _watcher_running = True
    _last_slot_seen  = None
    _watcher_thread  = threading.Thread(target=_slot_watcher_worker, daemon=True)
    _watcher_thread.start()
    logger.info("Slot watcher started.")


def stop_slot_watcher():
    global _watcher_running
    _watcher_running = False
    logger.info("Slot watcher stopped.")


def _slot_watcher_worker():
    """
    Polls every 30 seconds.
    Detects when slot changes → triggers auto-absent for completed slot.
    """
    global _last_slot_seen
    logger.info("Slot watcher thread running.")

    while _watcher_running:
        try:
            now          = datetime.now()
            current_slot = database.get_current_slot(now)
            current_label = current_slot["label"] if current_slot else None

            # Slot just ended: we had a slot before, now we don't (or different slot)
            if _last_slot_seen and _last_slot_seen != current_label:
                today = now.strftime("%Y-%m-%d")
                logger.info(f"Slot ended: {_last_slot_seen} — running auto-absent for {today}")
                count = database.mark_absent_for_slot(today, _last_slot_seen)
                logger.info(f"Auto-absent: {count} students marked Absent for {_last_slot_seen}")

            _last_slot_seen = current_label

        except Exception as e:
            logger.error(f"Slot watcher error: {e}")

        # Poll every 30 seconds
        time.sleep(30)


# ── CSV helper ─────────────────────────────────────────────────────────────────

def _write_csv(student_id, name, att_date, att_time, slot, status):
    _ensure_dir()
    path     = os.path.join(ATTENDANCE_DIR, f"attendance_{att_date}.csv")
    new_file = not os.path.isfile(path)
    with open(path, "a", newline="") as f:
        w = csv.writer(f)
        if new_file:
            w.writerow(["Student ID","Name","Date","Time","Slot","Status"])
        w.writerow([student_id, name, att_date, att_time, slot, status])


# ── Queries ────────────────────────────────────────────────────────────────────

def get_today_attendance() -> list:
    return database.get_attendance_by_date(date.today().strftime("%Y-%m-%d"))

def get_attendance_by_date(query_date) -> list:
    return database.get_attendance_by_date(query_date)

def get_all_attendance() -> list:
    return database.get_all_attendance()

def get_slots() -> list:
    return database.get_slots()


# ── Excel export ───────────────────────────────────────────────────────────────

def export_to_excel(output_path=None) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed."); return ""

    records = get_all_attendance()
    if not records: return ""

    _ensure_dir()
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(ATTENDANCE_DIR, f"attendance_export_{ts}.xlsx")

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Attendance"
    headers  = ["Student ID","Name","Date","Slot","Time","Status","Marked By"]
    hf       = Font(bold=True, color="FFFFFF")
    hfill    = PatternFill("solid", fgColor="0d3b66")
    alt      = PatternFill("solid", fgColor="e8f4fd")
    prs_fill = PatternFill("solid", fgColor="d4edda")  # green for Present
    abs_fill = PatternFill("solid", fgColor="f8d7da")  # red for Absent

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    for ri, rec in enumerate(records, 2):
        vals = [rec["student_id"], rec["name"], rec["date"],
                rec.get("slot",""), rec["time"],
                rec.get("status","Present"), rec.get("marked_by","auto")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(horizontal="center")
            if ri % 2 == 0: cell.fill = alt
        status = rec.get("status","Present")
        ws.cell(row=ri, column=6).fill = prs_fill if status=="Present" else abs_fill

    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = w + 4

    wb.save(output_path)
    logger.info(f"Exported: {output_path}")
    return output_path